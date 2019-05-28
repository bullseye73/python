# _*_ coding: utf-8 _*_

'''
    excel 다루기
    https://book.coalastudy.com/data-crawling/week-5/stage-2#worksheet
'''

import openpyxl, datetime, shutil

def srcFile():
    src = "weekreport.xlsx"
    fn = src.split('.')
    dst = "./data/" + fn[0] + "_" + str(datetime.datetime.now().timestamp()) + "." + fn[1]
    shutil.copy(src, dst, follow_symlinks=False)
    print(dst)
    return dst

def main():
    saveFile = srcFile()
    wb = openpyxl.load_workbook(saveFile)
    sheets = wb.sheetnames
    ws = wb[sheets[1]]  # 외환 sheet
    ws['f6'] = 7
    ws['f7'] = 2
    ws['f8'] = 2

    ws['f9'] = 7
    ws['f10'] = 2
    ws['f11'] = 2

    ws['f12'] = 7
    ws['f13'] = 2
    ws['f14'] = 2

    ws['g6'] = 6
    ws['g7'] = 3
    ws['g8'] = 3

    ws['g9'] = 6
    ws['g10'] = 3
    ws['g11'] = 3

    ws['g12'] = 6
    ws['g13'] = 3
    ws['g14'] = 3

    print(sheets)
    print(ws)

    wb.save(saveFile)

if __name__ == '__main__':
    main()
