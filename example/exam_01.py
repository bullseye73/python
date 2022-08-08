
import os
import re
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

#import xlsxwriter
g_fileName = "/Users/chang-hunjeong/workspace/python/example/data.json"
DEF_LINE = "============================================================="
#xlsFile = None

parse_data = {
             "PTT": "MICOM_PTT_PRESS:SHORT", 
             "IPC_SND_START" : "msgId : APP_REQUEST_RECOG_START_VRSVC", 
             "IPC_REV_START" : "msgId APP_REQUEST_RECOG_START_VRSVC", 
             "CONNECT_REQ" : "CloudWebsocketClient.cpp:028:connect", 
             "CONNECT_RES" : "Connected to websocket server requestId:", 
             "MIC_INPUT_START" : "MIC_INPUT_START", 
             "MIC_INPUT_END" : "MIC_INPUT_END", 
             "BOS" : "BEGIN_OF_SPEECH", 
             "EOS" : "END_OF_SPEECH", 
             "SERVER_SEND_START" : "Starting request frame", 
             "SERVER_SEND_END" : "Server recognizer requested embedded EOS", 
             "DICTATION_START" : "BPD", 
             "DICTATION_PARTIAL" : "PARTIAL", 
             "DICTATION_FINAL" : "FINAL", 
             "VR_RESULT" : "\[ART\]RESULT", 
             "IPC_SND_RESULT" : "msgId VRSVC_RESPOND_APP", 
             "IPC_REV_RESULT" : "msgId : VRSVC_RESPOND_APP", 
             "DISPLAY" : "msgId : \[VRSVC_RESPOND_APP\] msgData"}

# 입력된 리스트의 값으로 엑셀파일에 저장한다.
def openXlsFile():
    wb = Workbook()
    #ws = wb.create_sheet("parseData")
    ws = wb.active
    return ws

def saveXlsFile(_worksheet, _fileName):
    print(DEF_LINE)
    print(_fileName)
    print(DEF_LINE)
    _worksheet.save(_fileName)
    return True

def writeXlsFile(_worksheet, _inputData):
	rowTitle = []
	rowData = []
	
	
	for row in _inputData:
		if _worksheet.max_row == 1:
			rowTitle.append(row[0])
		if row[0] == 'CONNECT_RES':
			if _worksheet.max_row == 1 :
				rowTitle.append("TID")
			rowData.append(row[1])
			rowData.append(row[2])
			#print("TID : {}".format(row[2]))
		else:
			rowData.append(row[1])
			#print(row[1])
	if _worksheet.max_row == 1 :
		_worksheet.append(rowTitle)
	_worksheet.append(rowData)
	return True

# 로그가 들어가면 키값과 value가 리턴한다.
def match_log(_value):
	keyValue = ""
	sp = []
	result = re.compile('[a-z]+', re.I)

	for key in parse_data:
		#print("key:[{0}], value:[{1}]".format(key, parse_data[key]))
		#data.clear()
		sp = _value.split()
		result = re.search(parse_data[key], _value)
		
		if result != None:
			keyValue = key
			break
	
	if len(sp) <= 0: # is None:
		return keyValue
	if keyValue == '':
		return
	elif keyValue == 'CONNECT_RES':
		tid = re.sub("[\[\]]", "", sp[24])
		return keyValue, sp[2], tid

	return keyValue, sp[2]
		
def readFileName(path, _worksheet):
	row = 0
	result = []
	apData = []
	f = open(path, 'r', encoding='UTF8')
	while True:
		line = f.readline()
		if "----" in line:
			row = 0
			writeXlsFile(_worksheet, apData)
			apData.clear()
			print(DEF_LINE)
			continue
		if not line:
			#print("{}".format(result))
			break
		line = line.strip()  # 줄 끝의 줄 바꿈 문자를 제거한다.
		#print ("------- {} --------".format(line))
		result = match_log(line)
		if result != '':
			if result is not None:
				#print("{}".format(result))
				apData.append(result)
		row += 1
	f.close()
		

def main():
	wb = Workbook()
	ws = wb.active
	#ws.max_row
	
	fname, ext = os.path.splitext(g_fileName)
	#worksheet = openXlsFile()
	readFileName(g_fileName, ws)
	#saveXlsFile(ws, "{}.xlsx".format(fname))
	wb.save("{}.xlsx".format(fname))
	wb.close

if __name__ == '__main__':
    main()

