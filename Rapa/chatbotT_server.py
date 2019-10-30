#-*- coding:utf-8 -*-

import _thread, time
import codecs
import re
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import os, sys
from libs import diff_match_patch as dmp_module
from libs import commTime
from libs import rsapi

#url = 'http://192.168.219.115'
url = 'http://222.122.197.13'
prt = 8081

headers = {'Content-Type': 'application/json; charset=utf-8'}

def setThemeRow(ws, clr='NONE', isTitle=False):
    thin = Side(border_style="thin", color="FF000000")
    double = Side(border_style="double", color="FF000000")
    thick = Side(border_style="thick", color="FF000000")

    ncol = ws.max_column + 1

    for j in range(1, ncol):
        c = ws.cell(ws.max_row, j)
        c.font = Font(size=9, bold=False)

        if (clr != 'NONE') :
            if (isTitle == False) :
                c.font = Font(size=9, bold=False)
                if (j == 1 or j == ws.max_column):
                    if (j==ws.max_column):
                        if (c.value != 100):
                            clr = 'FFFF0000'
                    c.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    c.alignment = Alignment(horizontal="left", vertical="center")
            else:
                if (j == 1 or j == ws.max_column):
                    ws.column_dimensions[c.column_letter].width = 5
                else:
                    ws.column_dimensions[c.column_letter].width = 46

                c.font = Font(size=9, bold=True)
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border = Border(top=double, left=thin, right=thin, bottom=thick)

            c.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            c.fill = PatternFill(fill_type='solid', start_color=clr, end_color=clr)
        #else:
        #    c.font = Font(size=9, bold=False)


def cleanText (text):
    return re.sub('[;:/\-=?*’\'•«»<>+♦—()}]', '', text).rstrip().lstrip()

def recognitionRate(str1, str2):
    text1 = cleanText(str(str1))
    text2 = cleanText(str(str2))
    text1 = text1.upper().replace(" ", "")
    text2 = text2.upper().replace(" ", "")
    if str1 == None or str2 == None:
        return 0

    diffCount = countDiff(text1, text2)

    return diffCount

def countDiff(text1, text2):
    dmp = dmp_module.diff_match_patch()
    dmp.Diff_Timeout = 0.0
    diff = dmp.diff_main(text1, text2, False)

    # similarity
    common_text = sum([len(txt) for op, txt in diff if op == 0])
    text_length = max(len(text1), len(text2))
    sim = common_text / text_length
    return round(sim * 100, 1)

def compareData(org, ret):
    result = list()
    disCount = 0;
    try :
        for i in range(len(org)):
            print('compareData_ex : [{0}][{1}][{2}]'.format(i, org[i], ret[i]))
            if org[i] == None: #  and ret[i] == None:
                disCount += 1;
                result.append(0)
            elif i == 0 :
                result.append(0)
            elif i == 1:
                result.append(0)
            else:
                # 인식률(Recognition rate) 값
                result.append(recognitionRate(ret[i], org[i])) # (OCR 인식결과, 정답셋 글자수)
    except (IndexError, ValueError):
        print("index error [{0}],[{1}]".format(IndexError, ValueError))

    print('compareData_ex end disCount : [{0}]'.format(disCount))
    return result, disCount

def readTxtFile(fn):
    fname, ext = os.path.splitext(fn)
    rapi = rsapi.RsRequest(url, prt)
    #wb = openpyxl.load_workbook(fname + ".xlsx")
    wb = Workbook()
    ws = wb.active

    title = ['no', '질의', '정답지', '챗봇응답', '인식률']
    ws.append(title)
    setThemeRow(ws, 'FFF4F4F4', True)


    with codecs.open(fn, 'r', encoding="utf-8-sig") as f:
        idx = 0
        for line in f:
            xlsData = []
            idx += 1
            data = line.split('\\')
            xlsData.append(str(idx))
            xlsData.append(data[0])
            xlsData.append(data[1])
            body = { "user_id" : "claude", "msg" : re.sub('[?."\t]', '', data[0]) }
            #commTime.cp(re.sub('[?."]', '', data[0]))
            res = rapi.send('', body, headers)
            time.sleep(0.6)

            if res != None:
                #commTime.cp(res['msg_id'])
                for h in res['utterances']:
                    commTime.cp(data[0] + "\t" + h['text'])
                    xlsData.append(h['text'])
                    xlsData.append(recognitionRate(data[1], h['text']))

            ws.append(xlsData)
            setThemeRow(ws, 'FFFFFFFF', False)

    wb.save(fname + ".xlsx")
    wb.close()
    #return True

def request_chatbot_call (body, delay):
    count = 0
    while count == 0 :
        time.sleep(delay)


if __name__ == '__main__':
    if len(sys.argv) != 2:
        commTime.cp('input chatbot test file ~.txt')
        sys.exit()
    commTime.cp(sys.argv[1] + ' Start')
    #readTxtFile('testcase_name.txt')
    readTxtFile(sys.argv[1])
