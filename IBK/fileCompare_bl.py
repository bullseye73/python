#-*- coding: utf-8 -*-

import os, codecs
import openpyxl
import sys, re, time
import diff_match_patch as dmp_module

def makeFileName (str):
    fn = re.sub('[-=.#/?:$}]', '', str)
    now = time.localtime()
    s = "%04d%02d%02d_%02d%02d%02d" % (now.tm_year, now.tm_mon, now.tm_mday, now.tm_hour, now.tm_min, now.tm_sec)
    return fn + "_" + s + ".xlsx"

def makeFileName_ex (str):
    return str + ".xlsx"

def search(dir):
    files = os.listdir(dir)
    workbook = openpyxl.Workbook()
    workbook.remove(workbook['Sheet'])
    for file in files:
        fullFilename = os.path.join(dir, file)
        print("is file : {0} ".format(fullFilename))
        fname, ext = os.path.splitext(file)
        worksheet = workbook.create_sheet(fname)
        readTxtFile(worksheet, fullFilename)

    workbook.save(makeFileName_ex(fname))
    workbook.close()


def readXlsxFile(oriPath):
    wb = openpyxl.load_workbook(oriPath)
    ws = wb.active
    rows = ws.rows
    values = []

    for row in rows:
        cv = []
        for cell in row:
            cv.append(cell.value)
        values.append(cv)
    wb.close()
    return values

def compareData(org, ret):
    result = list()

    for i in range(len(org)):
        if org[i] == None and ret[i] == None:
            #print('[{0}][{1}][{2}]'.format(i, org[i], ret[i]))
            result.append('0')
        else:
            # 인식률(Recognition rate) 값
            #print('[{0}][{1}]'.format(org[i], ret[i]))
            result.append(recognitionRate(ret[i], org[i])) # (OCR 인식결과, 정답셋 글자수)
    return result

def recognitionRate(str1, str2):
    text1 = str(str1)
    text2 = str(str2)
    if str1 == None or str2 == None:
        return '0'

    diffCount = countDiff(text1, text2)
    return diffCount

def countDiff(text1, text2):
    dmp = dmp_module.diff_match_patch()
    dmp.Diff_Timeout = 0.0
    diff = dmp.diff_main(text1, text2, False)

    # similarity
    common_text = sum([len(txt) for op, txt in diff if op == 0])
    text_length = max(len(text1), len(text2))
    sim = common_text / text_length
    return sim * 100


def writeCompareResult(xlsxData, txtData, retFileName):
    if len(xlsxData) != len(txtData):
        print("Can not compare.")
        return

    workbook = openpyxl.Workbook()
    workbook.remove(workbook['Sheet'])
    worksheet = workbook.create_sheet(retFileName)

    nRows = len(xlsxData)
    #nCols = len(xlsxData[0])

    #print('[{0}][{1}]'.format(len(txtData), len(txtData[0])))
    for i in range(nRows):
        if i == 0:
            worksheet.append(xlsxData[i])
            continue
        else:
            worksheet.append(txtData[i])
            worksheet.append(xlsxData[i])
            worksheet.append(compareData(xlsxData[i], txtData[i]))
        #for j in range(1, nCols):
    os.remove(retFileName)
    workbook.save('ret_'+ retFileName)
    workbook.close()


def readTxtFile(ws, fn):
    #strType = fn.split('_', 1)
    with codecs.open(fn, 'r', encoding="utf-8-sig") as f:
        row = 1 #cell의 row, col의 값이 1부터임. 0이면 오류
        col = 1
        excepts = ['category', 'category number'] # 사용하지 않는 column

        for line in f:
            r = line.replace('\r\n', '').strip()
            rData = r.replace('"', '').split(',', 1)
            rlen = len(rData)

            if rlen <= 1:
                continue

            if 'set name' in rData[0].lower():
                row += 1
                col = 1
            elif rData[0].lower() in excepts:
                continue
            else:
                val = rData[rlen - 1].replace(",", "")
                if row == 2:
                    ws.cell(row=row-1, column=col).value = rData[0]

                ws.cell(row, col).value = val
                col += 1

def readTxtFile_ex(fn):
    workbook = openpyxl.Workbook()
    workbook.remove(workbook['Sheet'])
    fi = os.path.split(fn)
    fname, ext = os.path.splitext(fi[1])
    worksheet = workbook.create_sheet(fname)

    #strType = fn.split('_', 1)
    with codecs.open(fn, 'r', encoding="utf-8-sig") as f:
        row = 1 #cell의 row, col의 값이 1부터임. 0이면 오류
        col = 1
        excepts = ['category', 'category number'] # 사용하지 않는 column

        for line in f:
            r = line.replace('\r\n', '').strip()
            rData = r.replace('"', '').split(',', 1)
            rlen = len(rData)

            if rlen <= 1:
                continue

            if 'set name' in rData[0].lower():
                row += 1
                col = 1
            elif rData[0].lower() in excepts:
                continue
            else:
                #val = rData[rlen - 1].replace(",", "")
                val = rData[rlen - 1]
                if row == 2:
                    worksheet.cell(row=row-1, column=col).value = rData[0]

                worksheet.cell(row, col).value = val
                col += 1
    workbook.save(makeFileName_ex(fname))
    workbook.close()

def Usage():
    print ("Usage: input 2 file fullpath \n python fileCompare_bl.py [result txt fullpath name] [xlsx file fullpath name]")

def main():
    if len(sys.argv) != 3:
        Usage()
        sys.exit()
    fi = os.path.split(sys.argv[1])
    fname, ext = os.path.splitext(fi[1])

    readTxtFile_ex(sys.argv[1])

    txtRet = readXlsxFile(makeFileName_ex(fname))
    xlsxRet = readXlsxFile(sys.argv[2])
    writeCompareResult(xlsxRet, txtRet, makeFileName_ex(fname))


if __name__ == '__main__':
    main()