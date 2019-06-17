#-*- coding: utf-8 -*-

import os, codecs
import openpyxl
import sys, re, time

def makeFileName (str):
    fn = re.sub('[-=.#/?:$}]', '', str)
    now = time.localtime()
    s = "%04d%02d%02d_%02d%02d%02d" % (now.tm_year, now.tm_mon, now.tm_mday, now.tm_hour, now.tm_min, now.tm_sec)
    return fn + "_" + s + ".xlsx"

def search(dir):
    files = os.listdir(dir)
    workbook = openpyxl.Workbook()

    for file in files:
        fullFilename = os.path.join(dir, file)
        print("is file : {0} ".format(fullFilename))
        fname, ext = os.path.splitext(file)
        worksheet = workbook.create_sheet(fname)
        readTxtFile(worksheet, fullFilename)

    workbook.save(makeFileName(dir))
    workbook.close()

def readTxtFile(ws, fn):
    #strType = fn.split('_', 1)

    with codecs.open(fn, 'r', encoding="utf-8-sig") as f:
        row = 1
        col = 1

        for line in f:
            r = line.replace('\r\n', '').strip()
            rData = r.replace('"', '').split(',', 1)
            rlen = len(rData)

            if rlen <= 1:
                continue

            if 'set name' in rData[0].lower():
                row += 1
                col = 1
            else:
                val = rData[rlen - 1].replace(",", "")
                if row == 2:
                    ws.cell(row=row-1, column=col).value = rData[0]

                ws.cell(row, col).value = val
                col += 1

def Usage():
    print ("Usage: input folder [dir name]")

def main():
    if len(sys.argv) != 2:
        Usage()
        sys.exit()

    search(sys.argv[1])

if __name__ == '__main__':
    main()