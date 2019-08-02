#-*- coding: utf-8 -*-

import os, codecs
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import sys, re, time
from libs import diff_match_patch as dmp_module
import numpy as np

def makeFileName_ex (str):
    return str + ".xlsx"

def readXlsxFile(oriPath):
    wb = openpyxl.load_workbook(oriPath)
    ws = wb.active
    rows = ws.rows
    values = []

    for row in rows:
        cv = []
        for cell in row:
            cv.append(cell.value)
        values.append(cv)
    wb.close()
    return values

def compareData(org, ret):
    result = list()
    try :
        for i in range(len(org)):
            if org[i] == None : # or i == 0 or ret[i] == None:
                #print('[{0}][{1}][{2}]'.format(i, org[i], ret[i]))
                result.append(0)
            elif i == 0 :
                result.append(0)
            else:
                # 인식률(Recognition rate) 값
                print('compareData : [{0}][{1}]'.format(org[i], ret[i]))
                result.append(recognitionRate(ret[i], org[i])) # (OCR 인식결과, 정답셋 글자수)
    except (IndexError, ValueError):
        print("index error [{0}],[{1}]".format(IndexError, ValueError))

    return result


def compareData_ex(org, ret):
    result = list()
    disCount = 0;
    try :
        for i in range(len(org)):
            if org[i] == None: #  and ret[i] == None:
                #print('compareData_ex : [{0}][{1}][{2}]'.format(i, org[i], ret[i]))
                disCount += 1;
                result.append(0)
            elif i == 0 :
                result.append(0)
            else:
                # 인식률(Recognition rate) 값
                #print('compareData_ex : [{0}][{1}]'.format(org[i], ret[i]))
                result.append(recognitionRate(ret[i], org[i])) # (OCR 인식결과, 정답셋 글자수)
    except (IndexError, ValueError):
        print("index error [{0}],[{1}]".format(IndexError, ValueError))

    return result, disCount
'''
    인식율 결과 산출
    str1 : 정답 결과값
    str2 : OCR 인식값
'''
def recognitionRate(str1, str2):
    text1 = cleanText(str(str1))
    text2 = cleanText(str(str2))
    text1 = text1.upper().replace(" ", "")
    text2 = text2.upper().replace(" ", "")
    if str1 == None or str2 == None:
        return 0

    diffCount = countDiff(text1, text2)

    return diffCount

def countDiff(text1, text2):
    dmp = dmp_module.diff_match_patch()
    dmp.Diff_Timeout = 0.0
    diff = dmp.diff_main(text1, text2, False)

    # similarity
    common_text = sum([len(txt) for op, txt in diff if op == 0])
    text_length = max(len(text1), len(text2))
    sim = common_text / text_length
    return round(sim * 100, 1)

'''
    리스트 열의 평균값
    ls : 리스트 2차 행열
'''
def averageRecRate(ls):
    avrColSum = list()
    lRow = list()
    arrData = np.array(ls)
    totalCol = len(arrData[0])
    totalRow = len(arrData)

    for i in range(totalCol):
        lRow.clear()
        for j in range(totalRow):
            lRow.append(float(arrData[j][i]))
        #print("[{0}][{1}][{2}]".format(i, totalRow, arrData[j][i]))
        avrColSum.append(sum(lRow)/totalRow)
    #print(avrColSum)
    return avrColSum

def rateTypeCopy(lt, rtType=80):
    nlen = len(lt)
    rateData = list()

    for i in range(nlen) :
        if (lt[i] <= rtType) :
            rateData.append(0)
        else:
            rateData.append(lt[i])

    return rateData

def averageRow(lt, count):
    avr = 0

    if sum(lt) == 0:
        avr = sum(lt)
    else:
        avr = sum(lt) / (len(lt) - (count + 1))  # filename 필드 제외

    return avr

def setThemeRow(ws, clr='NONE', name='', isTitle=False):
    thin = Side(border_style="thin", color="FF000000")
    double = Side(border_style="double", color="FF000000")
    thick = Side(border_style="thick", color="FF000000")

    ncol = 0
    if (isTitle):
        ncol = 1
    else:
        ncol = 0

    for j in range(ws.max_column + ncol):
        c = ws.cell(ws.max_row, j + 1)

        if (clr != 'NONE') :
            if (isTitle == False) :
                if j == 0:
                    c.value = name
                    c.font = Font(size=9, bold=False)
                else:
                    c.value = round(c.value, 1)
                    c.font = Font(size=9, bold=True)
            else:
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border = Border(top=double, left=thin, right=thin, bottom=thick)

            c.fill = PatternFill(fill_type='solid', start_color=clr, end_color=clr)
        else:
            c.font = Font(size=9, bold=False)


'''
    OCR 결과와 정답 xlsx 과의 내용 비교
    xlsxData : 정답데이터
    txtData : OCR 인식 데이터
    retFileName : 결과 저장 파일명
    type : OCR 인식 타입 (BL, INVOICE, LC)
'''
def writeCompareResult(xlsxData, txtData, retFileName, type):
    if len(xlsxData) != len(txtData):
        print("Can not compare.")
        return

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = type

    nRows = len(xlsxData)
    comData = list()
    rt80Data = list()
    rt90Data = list()
    colAvr = list()
    col80Avr = list()
    col90Avr = list()

    for i in range(nRows):
        if i == 0:
            worksheet.append(xlsxData[i])
            # 구분색 추가, Title
            setThemeRow(worksheet, 'FF00fff2','' , True)
            continue
        else:
            #comData = compareData(xlsxData[i], txtData[i])
            comData, dc = compareData_ex(xlsxData[i], txtData[i])
            rt80Data = rateTypeCopy(comData, 80)
            rt90Data = rateTypeCopy(comData, 90)

            comData.append(averageRow(comData, dc))
            rt80Data.append(averageRow(rt80Data, dc))
            rt90Data.append(averageRow(rt90Data, dc))

            # 정답지
            worksheet.append(xlsxData[i])
            setThemeRow(worksheet)
            # 인식 결과
            worksheet.append(txtData[i])
            setThemeRow(worksheet)

            # 인식률
            worksheet.append(comData)
            setThemeRow(worksheet, 'fffff79d', '문자 단위 인식율')

            # 인식률 (80이하 버림.)
            worksheet.append(rt80Data)
            setThemeRow(worksheet, 'fffffabe', '필드 단위 인식율(기준값 80 이하=0)')

            # 인식률 (90이하 버림.)
            worksheet.append(rt90Data)
            setThemeRow(worksheet, 'fffffcdf', '필드 단위 인식율(기준값 90 이하=0)')

            colAvr.append(comData)
            col80Avr.append(rt80Data)
            col90Avr.append(rt90Data)

    worksheet.append(averageRecRate(colAvr))
    setThemeRow(worksheet, 'FFbbef6c', '문자 단위 인식율 평균')
    worksheet.append(averageRecRate(col80Avr))
    setThemeRow(worksheet, 'FFcbff7b', '필드 단위 인식율(기준값 80 이하=0) 평균')
    worksheet.append(averageRecRate(col90Avr))
    setThemeRow(worksheet, 'FFebff99', '필드 단위 인식율(기준값 90 이하=0) 평균')

    #os.remove(retFileName)
    workbook.save(type + "_ret_" + retFileName)     #
    workbook.close()

'''
    bl 결과 파일 읽어 xlsx로 저장
    fn : xlsx 파일명
'''
def readTxtFile(fn):
    fi = os.path.split(fn)
    fname, ext = os.path.splitext(fi[1])

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = fname

    #strType = fn.split('_', 1)
    with codecs.open(fn, 'r', encoding="utf-8-sig") as f:
        row = 1 #cell의 row, col의 값이 1부터임. 0이면 오류
        col = 1
        excepts = ['category', 'category number'] # 사용하지 않는 column

        for line in f:
            r = line.replace('\r\n', '').strip()
            rData = r.replace('"', '').split(',', 1)
            rlen = len(rData)

            if rlen <= 1:
                continue

            if 'set name' in rData[0].lower():
                row += 1
                col = 1
            elif rData[0].lower() in excepts:
                continue
            else:
                #val = cleanText(rData[rlen - 1])
                val = rData[rlen - 1]
                if row == 2:
                    c = worksheet.cell(row=row-1, column=col)
                    c.font = Font(size=9, bold=True)
                    c.fill = PatternFill(fill_type='solid', start_color='FFADFF2F', end_color='FFADFF2F')
                    c.value = rData[0]
                    #worksheet.cell(row=row-1, column=col).value = rData[0]


                worksheet.cell(row, col).value = val
                col += 1
    workbook.save(makeFileName_ex(fname))
    workbook.close()

def getOCRType(fn):
    lf = ["BL", "LC", "invoice", "RETIRE", "SM", "MI"]
    for i in lf:
        if i.lower() in fn.lower():
            return i

def Usage():
    print ("Usage: input 2 file fullpath \n python fileCompare_bl.py [result txt fullpath name] [xlsx file fullpath name]")

def cleanText (text):
    text = re.sub("[.,]", " ", text)
    return re.sub('[;:/\-=?*’\'•«»<>♦—()}]', '', text).rstrip().lstrip()

'''
BL test parameter
    ./result/BL_result_0710_01.txt ./excel/RA_BL_result_0702.xlsx
LC test parameter
    ./result/LC_result_0710_01.txt ./excel/RA_LC_result_0702.xlsx
INVOICE test parameter
    ./result/Invoice_result_0612.txt ./excel/RA_INVOICE_result_0702.xlsx  
'''
def main():
    if len(sys.argv) != 3:
        Usage()
        sys.exit()
    fi = os.path.split(sys.argv[1])
    fname, ext = os.path.splitext(fi[1])
    sType = getOCRType(fname).upper()

    print("Start read file : [{0}] Type : [{1}]".format(fname, ext))
    readTxtFile(sys.argv[1])

    txtRet = readXlsxFile(makeFileName_ex(fname))

    print("Compare with file : [{0}]".format(sys.argv[2]))
    xlsxRet = readXlsxFile(sys.argv[2])

    writeCompareResult(xlsxRet, txtRet, makeFileName_ex(fname), sType)
    print("Compare end output file : [{0}]", makeFileName_ex(fname))

if __name__ == '__main__':
    main()